import sqlite3

# connect to database
def createConnection(databaseLocation):
    conn = None
    try:
        conn = sqlite3.connect(databaseLocation)
    except Error as e:
        print(e)
    return conn

# get stats on player you want to perform a 20-man comp for
def findPlayer(conn, playerName):
        cur = conn.cursor()
        cur.execute("SELECT * FROM playerStats WHERE Player=?", (playerName,))

        # in the case of duplicate names, just take the player with the highest scoring average
        # this should be the player you wanted ~almost~ every time
        player = cur.fetchone()
        if player == None:
            print(str(playerName) + " not found, please retry.")
        else:
            return player


# find the 20 players closest to the player in question using the formula below
# formula: ((|player1 PPG - player2 PPG|) * 0.4) + (|player1 APG - player2 APG|) * 0.25) + (|player1 RPG - player2 RPG|) * 0.25) + (|player1 FGP - player2 FGP|) * 0.1))
def findSimilarPlayers(conn, player):
    playerName = player[1]
    playerPPG = player[21]
    playerAPG = player[18]
    playerRPG = player[17]
    playerFGP = player[6]

    cur = conn.cursor()
    cur.execute("SELECT * FROM playerStats WHERE Player!=?", (playerName,))
    allPlayers = cur.fetchall()

    # add all players to dictionary where key = comp score and value = list of player info
    # then find the 20 most similar players
    # reduce dictionary to those 20 players (will be basis of our comparision)
    # return sorted dictionary
    allPlayersDict = {}
    for player in allPlayers:
        compScore = ((abs(playerPPG - player[21]) * 0.4) + (abs(playerAPG - player[18]) * 0.25) + (abs(playerRPG - player[17]) * 0.25) + (abs(playerFGP - player[6]) * 0.1))
        allPlayersDict[compScore] = player
    
    compPlayers = {}
    sortedKeys = sorted(allPlayersDict)
    sortedKeys = sortedKeys[:20]    # <------- This guy controls how many players are in the comp
    for key in sortedKeys:
        compPlayers[key] = allPlayersDict.get(key)   
    return compPlayers
    
        

# open comps excel worksheet and create a new sheet for current player
# write the data for the 20 player comps for that player
from openpyxl import load_workbook
# from openpyxl.utils import get_column_letter
def outputCompsToExcel(excelName, player, compDict):
    wb = load_workbook(filename = excelName)
    ws = wb.create_sheet(title=player[1])

    # write header of excel sheet
    headerList = ["Null", "ID", "Player", "Active", "GP", "MPG", "FGM", "FGA", "FGP", "TPM", "TPA", "TPP", "FTM", "FTA", "FTP", "TOV", "PF", "ORB", "DRB", "RPG", "APG", "SPG", "BPG", "PPG"]
    for c in range(1, len(headerList)):
        cellref = ws.cell(row = 1, column = c)
        cellref.value = headerList[c]

    # add player in question stats to top of excel
    for c in range(1, len(player) + 1):
        cellref = ws.cell(row = 2, column = c)
        cellref.value = player[c - 1]

    # change keys of dicitionary to ints so it can be iterated through 
    # (right now the keys are floats, this prevents us from assigning them to a row number when populating the excel)
    finalDict = {}
    count = 1
    for key in compDict:
        tempList = compDict.get(key)
        finalDict[count] = tempList
        count += 1

    # write player data for 20 player comps
    for r in range(3, len(finalDict) + 3):
        tempList = finalDict.get(r - 2)
        # print len(tempList)
        for c in range(1, len(tempList) + 1):
            cellref = ws.cell(row = r, column = c)
            cellref.value = tempList[c - 1]

    wb.save(filename = excelName)
    return finalDict

# returns a list of the 20 players in the comp
def getPlayerList(compDict):
    playerList = []
    for key in compDict:
        playerList.append(compDict.get(key)[1])
    return playerList

# function to get a player's ID by providing their name
from nba_api.stats.static import players
def getPlayerID(playerName):
    try:
        playerList = players.find_players_by_full_name(playerName)
        print(playerList[0])
        playerID = playerList[0].get("id")
        print(playerID)
        return playerID
    except:
        print("player not found")
        return 0

# function which writes a players complete game log to a CSV
# we can then take that CSV and parse it by season to get the season averages
# final result is a dictionary which contains the season averages of each player in our comp
import pandas as pd
from nba_api.stats.endpoints import playergamelog
from nba_api.stats.library.parameters import SeasonAll
import os
def getCareerGameStats(playerName, playerID):
    if playerID == 0:
        return None
    # pd.set_option("display.max_rows", None, "display.max_columns", None)

    completeGamelog = playergamelog.PlayerGameLog(player_id=playerID, season = SeasonAll.all)
    completeCareerDataFrame = completeGamelog.get_data_frames()[0]
    completeCareerDataFrame

    relativePath = "./"
    filename = relativePath + playerName + '.csv'
    completeCareerDataFrame.to_csv(filename)

    # get a list of each season the player has played (and a count of seasons)
    # to do so, parse through CSV and get unique season IDs
    # reverse this list so it becomes chronological
    # then get the season averages of each stat category for that season
    # store this as a dictionary
    # then graph the data
    seasonsPlayed = []
    seasonCount = 0

    df = pd.read_csv(filename)
    seasonCol = df.SEASON_ID
    for season in seasonCol:
        if season not in seasonsPlayed:
            seasonsPlayed.append(season)
            seasonCount += 1

    seasonsPlayed.reverse()
    seasonAverages = {}
    count = 0
    for season in seasonsPlayed:
        count += 1
        totalPoints = 0
        totalAssists = 0
        totalRebounds = 0
        gameCount = 0
        for index, row in df.iterrows():
            if row["SEASON_ID"] == season:
                totalPoints += row["PTS"]
                totalAssists += row["AST"]
                totalRebounds += row["REB"]
                gameCount += 1
        PTS = round(totalPoints / gameCount, 1)
        AST = round(totalAssists / gameCount, 1)
        REB = round(totalRebounds / gameCount, 1)
        tempList = [season, PTS, AST, REB]
        seasonAverages[count] = tempList

    seasons = seasonAverages.keys()
    points = []
    for key in seasons:
        points.append(seasonAverages.get(key)[1])

    # delete the CSV and return the dictionary
    os.remove(filename)
    return seasonAverages
    # print(seasons)
    # print(points)
    # plt.plot(seasons, points, label="Points")
    # plt.show()

def createGraphingDictionary(statDict):
    # get the longest career in the comp to determine the size of the dictionary
    longestCareer = 0
    players = statDict.keys()
    for player in players:
        tempDict = statDict.get(player)
        seasons = tempDict.keys()
        if len(seasons) > longestCareer:
            longestCareer = len(seasons)
        
    # reorganize dictionary so: {key = season #, val = [PTS player 1 in that season, PTS player 2 in that season, ...] }
    # start by creating empty dictionary with seasons from 1 to longest career and empty list for each val
    graphingDict = {}
    for season in range(1, longestCareer + 1):
        tempList = []
        graphingDict[season] = tempList
    # print(graphingDict)

    # now populate graphing dictionary & return
    for player in players:
        tempDict = statDict.get(player)
        seasons = tempDict.keys()
        for season in seasons:
            graphingDict[season].append(tempDict.get(season)[1])

    print(graphingDict)
    return graphingDict
            

# function which averages the statistics of each player in the comp and graphs the results
import matplotlib.pyplot as plt
def graphResults(graphingDict, playerName):
    keys = graphingDict.keys()
    x = []
    y = []
    for key in keys:
        x.append(key)
        pointList = graphingDict.get(key)
        avg = round(sum(pointList) / len(pointList), 1)
        y.append(avg)

    plt.plot(x, y)
    plt.xlabel("Season Number")
    plt.ylabel("PPG")
    plt.title("20 Man Comp: " + playerName)
    plt.show()



def main():
    # user input to generate desired data
    playerName = "Charles Barkley"
    databaseLocation = "../nba/stats.db"
    excelName = "comps.xlsx"

    # connect to database, find player, find comps, output comps to excel, generate player list
    conn = createConnection(databaseLocation)
    player = findPlayer(conn, playerName)
    print(player)
    compPlayers = findSimilarPlayers(conn, player)
    compDict = outputCompsToExcel(excelName, player, compPlayers)
    playerList = getPlayerList(compDict)

    # create dictionary of the season stats of each player
    statDict = {}
    for player in playerList:
        playerID = getPlayerID(player)
        if playerID != 0:
            individualPlayerStatsDict = getCareerGameStats(player, playerID)
            statDict[player] = individualPlayerStatsDict
    
    print(statDict)
    statDictSize = statDict.keys()
    print()
    print(len(statDictSize))
    print()
    
    # reorganize dictionary for easier graphing
    graphingDict = createGraphingDictionary(statDict)
    graphResults(graphingDict, playerName)


main()